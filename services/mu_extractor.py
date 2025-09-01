import pandas as pd
import openpyxl
import re
import hashlib
import zipfile
import os
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as XLImage
from typing import Iterable
import psycopg
from psycopg.rows import tuple_row, dict_row
from decimal import Decimal, InvalidOperation
from pathlib import Path



IMAGES_OUT_DIR = r'C:\Users\Dell\OneDrive - Xanita\Projects\mu_sheets\extracted_images'


# DB_DSN   = "postgresql://postgres:MattB01@localhost:5432/xanita-app"
DB_DSN = os.environ["DATABASE_URL"]
PROCESS  = "mu_extractor"
ETL_VERSION = 1
EXTRACT_IMAGES = False

df_jobs = pd.DataFrame(columns=['Job ID', 'Name'])
df_dims = pd.DataFrame(columns=['Job ID', 'Width', 'Height', 'Depth'])
df_boards = pd.DataFrame(columns=['Job ID', 'XB Type', 'Thickness', 'Size', 'Units Up'])


errors = []

job_rows = []
dims_rows = []
board_rows = []
image_rows = []

NULL_TOKENS = {"", "none", "null", "n/a", "na", "-", "—"}


def load_state_db(conn, process=PROCESS):
    with conn.cursor(row_factory=tuple_row) as cur:
        cur.execute(
            "SELECT etl_version, last_mtime, last_path FROM etl_state WHERE process=%s",
            (process,),
        )
        row = cur.fetchone()
    if not row or row[0] != ETL_VERSION:
        return {"etl_version": ETL_VERSION, "last_mtime": 0, "last_path": ""}
    return {"etl_version": row[0], "last_mtime": int(row[1] or 0), "last_path": row[2]}

def file_mtime(path_str: str) -> int:
    try:
        return int(os.path.getmtime(path_str))
    except (FileNotFoundError, PermissionError):
        return 0
    
def ensure_state_table(conn):
    with conn.cursor() as cur:
        cur.execute("""
        CREATE TABLE IF NOT EXISTS etl_state (
          process      text PRIMARY KEY,
          etl_version  int     NOT NULL,
          last_mtime   bigint  NOT NULL,
          last_path    text    NOT NULL,
          updated_at   timestamptz NOT NULL DEFAULT now()
        );
        """)
    conn.commit()

def save_state_db(conn, last_mtime, last_path, process=PROCESS):
    with conn.cursor() as cur:
        cur.execute("""
          INSERT INTO etl_state(process, etl_version, last_mtime, last_path)
          VALUES (%s,%s,%s,%s)
          ON CONFLICT (process) DO UPDATE
            SET etl_version=EXCLUDED.etl_version,
                last_mtime =EXCLUDED.last_mtime,
                last_path  =EXCLUDED.last_path,
                updated_at =now();
        """, (process, ETL_VERSION, int(last_mtime), last_path))
    conn.commit()

def to_decimal(val):
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in NULL_TOKENS:
        return None
    # normalize decimal comma and strip non-numeric (keep digits, dot, minus)
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9\.\-]", "", s)
    if s in ("", ".", "-"):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None

def to_int(val):
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in NULL_TOKENS:
        return None
    # strip non-digits/minus
    s = re.sub(r"[^0-9\-]", "", s)
    if s in ("", "-"):
        return None
    try:
        return int(s)
    except ValueError:
        return None

def unique_id(filename):
    # Create the base string
    base = filename
    
    # Compute SHA1 hash
    h = hashlib.sha1(base.encode("utf-8")).hexdigest()
    
    # Take only the first 6 chars (enough for uniqueness in small datasets)
    short_hash = h[:6]
    
    return f"{short_hash}"

def extract_jobs(vals, filename):
    # Flatten to clean strings
    txts = [str(v).strip() for v in vals if v is not None and str(v).strip() != ""]
    lower = [t.lower() for t in txts]

    job_no = None
    job_name = None

    # --- Try inline "Job no: 12345" / "Job number 12345" / "Job #12345"
    for t in txts:
        m = re.search(r'(?i)\bjob\s*(?:no\.?|number|#)\s*:?[\s\-]*([0-9]{3,})\b', t)
        if m:
            job_no = m.group(1)
            break

    # --- Try label followed by next cell
    if job_no is None:
        for i, t in enumerate(lower):
            if re.fullmatch(r'job\s*(?:no\.?|number|#)\s*:?', t):
                if i + 1 < len(txts):
                    digits = re.sub(r'\D', '', txts[i+1])
                    if digits:
                        job_no = digits
                        break

    # --- Job name: inline or label + next cell
    for i, t in enumerate(lower):
        if re.fullmatch(r'job\s*name\s*:?', t) or re.fullmatch(r'project\s*name\s*:?', t):
            if i + 1 < len(txts):
                job_name = txts[i+1]
                break
        m = re.search(r'(?i)\bjob\s*name\s*:?\s*(.+)$', txts[i])
        if m and m.group(1).strip():
            job_name = m.group(1).strip()
            break

    # --- Fallbacks from the file path
    p = Path(filename)
    # job folder e.g. ".../Job10023-Client-Thing/.../sheet.xlsx"
    job_folder = next((part for part in p.parts if re.match(r'(?i)^job\d+', part)), None)

    if job_no is None:
        # Try to pull digits from folder or filename
        m = re.search(r'(?i)job\s*0*([0-9]{3,})', str(p))
        if m:
            job_no = m.group(1)

    if job_name is None:
        # Use the job folder name as a decent fallback
        job_name = job_folder or p.parent.name

    if job_no is None:
        # still no job number -> signal the caller to skip
        return None

    # Keep your current uid scheme (file-hash) to avoid schema churn today
    uid = unique_id(filename)

    return {"ID": uid, "Job ID": job_no, "Name": job_name}


def extract_board(uid, vals):

    rows = []
    t = next((i for i, v in enumerate(vals) if v is not None and "BOARD REQUIRED" in str(v).upper()), None)
    if t is None:
        return rows  # no board section, bail gracefully

    # find end anchor after t
    end_needles = ["HARDWARE", "ELECTRICAL", "OUTSOURCED", "OUTSOURCING", "FINISHING"]
    e = next((i for i, v in enumerate(vals[t+1:], t+1)
              if v is not None and any(n in str(v).upper() for n in end_needles)), None)
    if e is None:
        e = len(vals)

    start = t+54
    for pos in range(t,e):
    # start = t+17
        while start<e:
            row = {
                # 'Job ID': job_no,
                'ID': uid,
                'XB Type': vals[start],
                'Thickness (mm)': re.sub(r"\s*mm$", "", str(vals[start+1]).strip(), flags=re.IGNORECASE),
                'Size': vals[start+2],
                'Units Up': vals[start+4]
            }
            rows.append(row)
            start = start+18

    return rows
    


def extract_dims(uid, vals):

    try:
        d = vals.index("Dims")
        dims = vals[d+1]
    except ValueError:  # "Dims" not found
        dims = 0

    if not isinstance(dims, str) or not dims.strip():
        # return {"Job ID": job_id, "Width": 0, "Height": 0, "Depth": 0}
        return {"ID": uid, "Width": 0, "Height": 0, "Depth": 0}
    s=dims.lower()
    s = re.sub(r"\s+", " ", s)
    pattern = r"(\d+(?:[.,]\d+)?)\s*(?:mm|cm|m)?\s*([whd])?"
    matches = re.findall(pattern, s)

    dims_clean = {"w":None, "h": None, "d":None}
    nums = []

    for num, label in matches:
        num = float(num.replace(",", "."))
        unit_match = re.search(rf"{num}[.,]?\d*\s*(mm|cm|m)", s)
        unit = unit_match.group(1) if unit_match else "mm"
        if unit == "cm":
            num *= 10
        elif unit == "m":
            num *= 1000
        val = int(round(num))

        if label in ("w", "h", "d"):
            dims_clean[label] = val
        else:
            nums.append(val)

    unlabeled = [k for k,v in dims_clean.items() if v is None]
    for k,v in zip(unlabeled, nums):
        dims_clean[k] = v

    return {
        # "Job ID":job_id,
        "ID": uid,
        "Width": dims_clean["w"],
        "Height": dims_clean["h"],
        "Depth": dims_clean["d"]
        # "raw_text": text
    }

# def shrink_image(path, max_width, max_height):
#     """
#     Resize image proportionally so it fits within max_width × max_height.
#     Returns a path to a temporary resized copy.
#     """
#     with PILImage.open(path) as im:
#         im.thumbnail((max_width, max_height))  # in-place resize, keeps aspect ratio
#         # new_path = path.replace(".","_", 1)  # e.g., mypic_small.png
#         im.save(path)
#         return path


# def extract_images_for_file(excel_path: str, uid: str, output_dir: str = IMAGES_OUT_DIR):
#     """
#     Extracts /xl/media/* from a single .xlsx, renames to {uid}__{idx}.{ext}, returns list of saved paths.
#     """
#     os.makedirs(output_dir, exist_ok=True)
#     saved = []
#     with zipfile.ZipFile(excel_path, 'r') as z:
#         media = [f for f in z.namelist() if f.startswith("xl/media/")]
#         for idx, m in enumerate(media, start=1):
#             ext = os.path.splitext(m)[1].lower()  # keep original ext
#             out_path = os.path.join(output_dir, f"{uid}__{idx}{ext}")
#             with z.open(m) as src, open(out_path, "wb") as dst:
#                 dst.write(src.read())
#             saved.append(out_path)
#     return saved




def write_mu_to_postgres_conn(conn, job_rows, dims_rows, board_rows):
    uids = {r["ID"] for r in job_rows} | {r["ID"] for r in dims_rows} | {r["ID"] for r in board_rows}
    uids = list(uids)

    with conn.cursor(row_factory=tuple_row) as cur:
        sql_job = """
        INSERT INTO mu_jobs (uid, job_id, job_name)
        VALUES (%s, %s, %s)
        ON CONFLICT (uid) DO UPDATE
              SET job_id = EXCLUDED.job_id,
                  job_name = EXCLUDED.job_name
        """
        data_jobs = []
        for r in job_rows:
            uid = r["ID"]
            job_id = str(r["Job ID"]).lower()
            job_name = r["Name"]
            data_jobs.append((uid, job_id, job_name))

        sql_dims = """
        INSERT INTO mu_dimensions (uid, width_mm, height_mm, depth_mm)
        VALUES (%s,%s,%s,%s)
        ON CONFLICT (uid) DO UPDATE
              SET width_mm  = EXCLUDED.width_mm,
                  height_mm = EXCLUDED.height_mm,
                  depth_mm  = EXCLUDED.depth_mm
        """
        data_dims = []
        for r in dims_rows:
            uid = r["ID"]
            width = to_int(r["Width"])
            height = to_int(r["Height"])
            depth = to_int(r["Depth"])
            data_dims.append((uid, width, height, depth))

        if uids:
            cur.execute("DELETE FROM mu_boards WHERE uid = ANY(%s)", (uids,))

        sql_boards = """
        INSERT INTO mu_boards (uid, xb_type, thickness_mm, size_text, units_up)
        VALUES (%s,%s,%s,%s,%s)
        """
        data_boards = []
        for r in board_rows:
            uid = r["ID"]
            xb_type = r["XB Type"]
            thickness = to_decimal(r["Thickness (mm)"])  # numeric
            size = r["Size"]
            units = to_decimal(r["Units Up"])
            data_boards.append((uid, xb_type, thickness, size, units))

        if data_jobs:
            cur.executemany(sql_job, data_jobs)
        if data_dims:
            cur.executemany(sql_dims, data_dims)
        if data_boards:
            cur.executemany(sql_boards, data_boards)

    conn.commit()

# ---------- BUILD CANDIDATES FROM WATERMARK ----------
with psycopg.connect(DB_DSN, row_factory=dict_row) as conn:
    # Keep your existing state table to process incrementally
    ensure_state_table(conn)
    st = load_state_db(conn)  # {"etl_version":..., "last_mtime":..., "last_path":...}
    last_m, last_uid = st["last_mtime"], st["last_path"]  # we’ll reuse last_path as "last_uid"

    # Pull locations from DB with their updated_at as an epoch (bigint), ordered + tie-broken by uid
    with conn.cursor() as cur:
        cur.execute("""
            SELECT
              uid,
              job_name,
              filepath,
              EXTRACT(EPOCH FROM updated_at)::bigint AS updated_epoch
            FROM mu_locations
            ORDER BY updated_at, uid
        """)
        rows = cur.fetchall()

    # Build incremental candidate list using the same watermark idea:
    candidates: list[tuple[int, str, str]] = []  # (updated_epoch, filepath, loc_uid)
    for r in rows:
        m = int(r["updated_epoch"])
        loc_uid = r["uid"]
        fp = r["filepath"]
        if (m > last_m) or (m == last_m and loc_uid > last_uid):
            if os.path.isfile(fp):  # only process files that exist
                candidates.append((m, fp, loc_uid))

    if not candidates:
        print("MU extractor: nothing to do.")
    else:
        candidates.sort(key=lambda t: (t[0], t[2]))  # (updated_epoch, uid)
        print(f"MU extractor: {len(candidates)} file(s) to process.")

        # ---------- PROCESS ONE FILE AT A TIME + ADVANCE WATERMARK ----------
        for mtime, file, loc_uid in candidates:
            try:
                wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
                ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active  # fallback
            except Exception as e:
                print(f"[SKIP] {file} load error: {e}")
                # If you want: advance watermark to avoid re-trying the same broken file forever
                save_state_db(conn, mtime, loc_uid)
                continue

            # flatten fast
            vals = []
            for row in ws.iter_rows(values_only=True):
                vals.extend(row)
            wb.close()

            job_row = extract_jobs(vals, file)
            if not job_row:
                print(f"[SKIP] {file}: Job no / Job Name not found")
                save_state_db(conn, mtime, loc_uid)  # advance so we don't re-try endlessly
                continue

            uid = job_row["ID"]  # your current ID = hash(filename) — OK to keep for now

            boards = extract_board(uid, vals) or []
            boards = [r for r in boards if all((v is not None) and (str(v).strip() != "") for v in r.values())]
            dims_row = extract_dims(uid, vals)

            # Optional images
            # Optional images (now disabled)
            if EXTRACT_IMAGES:
                try:
                    # if is_zip_excel(file):  # only if you had this helper; otherwise skip this line too
                    for pth in extract_images_for_file(file, uid, IMAGES_OUT_DIR):
                        image_rows.append({"ID": uid, "ImagePath": pth})
                except Exception as e:
                    print(f"[IMG] {file}: {e}")


            # keep for Excel export (optional)
            job_rows.append(job_row)
            dims_rows.append(dims_row)
            if boards:
                board_rows.extend(boards)

            # write THIS file to Postgres
            try:
                write_mu_to_postgres_conn(conn, [job_row], [dims_row], boards)
            except Exception as e:
                print(f"[DB-FAIL] {file}: {e}")
                # do NOT advance watermark on DB write failure, so we retry next run
                continue

            # advance watermark AFTER success (use loc_uid as tiebreaker)
            save_state_db(conn, mtime, loc_uid)

# ---------- OPTIONAL: build DFs for Excel export ----------
# df_jobs   = pd.DataFrame(job_rows, columns=['ID','Job ID','Name'])
# df_boards = pd.DataFrame(board_rows, columns=['ID','XB Type','Thickness (mm)','Size','Units Up']).dropna()
# df_dims   = pd.DataFrame(dims_rows, columns=['ID', 'Width', 'Height', 'Depth'])
# df_images = pd.DataFrame(image_rows, columns=['ID', 'ImagePath'])